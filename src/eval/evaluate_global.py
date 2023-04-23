import json
import os
import numpy as np
import openpyxl
import torch
from openpyxl import Workbook
from sklearn.metrics import classification_report
from torch import nn
from dotenv import load_dotenv

from src.data.data_FL import MyDataset, val_transform
from src.eval.database import Database
from src.modules import models
from src.modules.settings import DATASET

BATCH_SIZE = 16
RESULTS_FILE = 'All_Scores_Global.xlsx'
DB_FILE = 'results.db'
DB = Database(DB_FILE)
load_dotenv()


def predict(model, loader):
    model.eval()
    y_gt = []
    y_preds = []
    with torch.no_grad():
        for batch_idx, sample_batched in enumerate(loader):
            X = sample_batched[0].type(torch.cuda.FloatTensor)
            # img_labels = [DATASET.label_map[label] for label in sample_batched[1]]
            # y = torch.tensor(img_labels, dtype=torch.long).to('cuda')
            y = sample_batched[1].type(torch.cuda.LongTensor)
            y_gt.append(y.cpu())
            y_pred = model(X)
            y_pred = np.argmax(y_pred.cpu(), axis=1)
            y_preds.append(y_pred)
    return np.concatenate(y_gt), np.concatenate(y_preds)


def calculate_scores(y, predictions):
    """
    Calculate scores using scikit-learn

    Parameters:
        y: ground truths
        predictions: predictions

    Returns:
        scores: dictionary containing accuracy, F1-score, precision, recall, and class-based F1-scores
    """

    # Calculate class-based F1-scores
    class_report = classification_report(y, predictions, output_dict=True, zero_division=True)
    class_f1_scores = {DATASET.classes[int(k)]: v['f1-score'] for k, v in class_report.items() if k.isnumeric()}

    # Calculate class-based accuracies
    unique_classes = set(y)
    class_accuracies = {}
    for cls in unique_classes:
        correct = 0
        total = 0
        for true_label, pred_label in zip(y, predictions):
            if true_label == cls:
                total += 1
                if true_label == pred_label:
                    correct += 1
        class_accuracy = correct / total if total > 0 else 0
        class_accuracies[DATASET.classes[cls]] = class_accuracy

    # Create dictionary to store the scores
    scores = {
        'accuracy': class_report['accuracy'],
        'f1': class_report['weighted avg']['f1-score'],
        'precision': class_report['weighted avg']['precision'],
        'recall': class_report['weighted avg']['recall'],
        'class_f1_scores': class_f1_scores,
        'class_accuracies': class_accuracies
    }

    return scores


def prepare_excel(filename: str):
    # Create a new workbook object
    workbook = Workbook()
    # Add a sheet to the workbook
    sheet = workbook.active
    sheet.title = 'Val'
    workbook.create_sheet('Test')

    workbook[workbook.sheetnames[0]]['B1'] = 'Accuracy'
    workbook[workbook.sheetnames[0]]['C1'] = 'F1'
    workbook[workbook.sheetnames[0]]['D1'] = 'Precision'
    workbook[workbook.sheetnames[0]]['E1'] = 'Recall'

    workbook[workbook.sheetnames[1]]['B1'] = 'Accuracy'
    workbook[workbook.sheetnames[1]]['C1'] = 'F1'
    workbook[workbook.sheetnames[1]]['D1'] = 'Precision'
    workbook[workbook.sheetnames[1]]['E1'] = 'Recall'
    # Save the workbook to a file
    workbook.save(filename)


def generate_summary(model_path):
    """
    generate model summary and save it an excel sheet

    Parameters:
        None

    Returns:
        None
    """
    models_list = [model_path]

    model = models.get_model(DATASET.num_classes)
    device = torch.device('cuda:0')
    model = nn.DataParallel(model)
    model = model.cuda()
    model = model.to(device)

    prepare_excel(RESULTS_FILE)
    wb = openpyxl.load_workbook(RESULTS_FILE)
    i = 1
    shift = 1
    for name in models_list:
        test_ds, val_ds = get_dataset()

        valid_loader = torch.utils.data.DataLoader(val_ds, batch_size=BATCH_SIZE, shuffle=False, num_workers=0,
                                                   pin_memory=True)
        test_loader = torch.utils.data.DataLoader(test_ds, batch_size=BATCH_SIZE, shuffle=False, num_workers=0,
                                                  pin_memory=True)

        model.load_state_dict(torch.load(name))
        model.eval()

        worksheets = [[wb[wb.sheetnames[0]], 'validation'], [wb[wb.sheetnames[1]], 'test']]
        for worksheet in worksheets:
            ws = worksheet[0]
            y, predictions = predict(model, valid_loader if worksheet[1] == 'validation' else test_loader)
            scores = calculate_scores(y, predictions)
            DB.insert_result(os.path.basename(name), scores['accuracy'], scores['f1'], scores['precision'], scores['recall'],
                             worksheet[1], os.getenv('MODEL'), json.dumps(scores['class_f1_scores']),
                             json.dumps(scores['class_accuracies']))
            cl = 'A' + str(i + shift)
            ws[cl] = name
            cl = 'B' + str(i + shift)
            ws[cl] = scores['accuracy']
            cl = 'C' + str(i + shift)
            ws[cl] = scores['f1']
            cl = 'D' + str(i + shift)
            ws[cl] = scores['precision']
            cl = 'E' + str(i + shift)
            ws[cl] = scores['recall']
            cl = 'F' + str(i + shift)

    wb.save(RESULTS_FILE)


def get_dataset():
    images, labels, test, validation = DATASET.get_global_test_data()

    img_test = images[test]
    lbl_test = labels[test]

    img_validation = images[validation]
    lbl_validation = labels[validation]

    test_ds = MyDataset(img_test, lbl_test, val_transform)
    val_ds = MyDataset(img_validation, lbl_validation, val_transform)

    return test_ds, val_ds


if __name__ == '__main__':
    generate_summary(os.getenv('MODEL_FILE'))
