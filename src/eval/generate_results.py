import glob

import numpy as np
import openpyxl
import torch
import torch.nn as nn
from openpyxl import Workbook
from sklearn.metrics import precision_recall_fscore_support

from src.data.data_FL import val_transform, MyDataset
from src.modules import models

# Compute precision, recall, F-measure and support for each class.
# The precision is intuitively the ability of the classifier not to label as positive a sample that is negative.
# The recall is intuitively the ability of the classifier to find all the positive samples.
# The F-beta score can be interpreted as a weighted harmonic mean of the precision and recall.
# The F-beta score weights recall more than precision by a factor of beta.
# beta == 1.0 means recall and precision are equally important.
# The support is the number of occurrences of each class in y_true.
from src.modules.settings import DATASET

num_classes = 5
batchSize = 16
data_path = '../../dataset/'
name = 'FedPerl'
RESULTS_FILE = 'All_Scores.xlsx'
DATASET.clients_path = '../../dataset/clients/'

def predict(model, loader):
    model.eval()
    y_gt = []
    y_preds = []
    with torch.no_grad():
        for batch_idx, sample_batched in enumerate(loader):
            X = sample_batched[0].type(torch.cuda.FloatTensor)
            # img_labels = [LABEL_MAP[label] for label in sample_batched[1]]
            # y = torch.tensor(img_labels, dtype=torch.long).to('cuda')
            y = sample_batched[1].type(torch.cuda.LongTensor)
            y_gt.append(y.cpu())
            y_pred = model(X)
            y_pred = np.argmax(y_pred.cpu(), axis=1)
            y_preds.append(y_pred)
    return np.concatenate(y_gt), np.concatenate(y_preds)


def calculate_scores(y, predictions):
    """
    calculate scores using skitlearn

    Parameters:
        y: ground truths
        predictions: predictions

    Returns:
        scores
    """
    acc = np.mean(np.equal(predictions, y))
    prf3 = precision_recall_fscore_support(y, predictions, average='weighted', zero_division=1)

    return acc, prf3[0], prf3[1], prf3[2]


def prepare_excel(filename: str):
    # Create a new workbook object
    workbook = Workbook()
    # Add a sheet to the workbook
    sheet = workbook.active
    sheet.title = 'Val'
    workbook.create_sheet('Test')

    workbook[workbook.sheetnames[0]]['B1'] = 'Accuracy'
    workbook[workbook.sheetnames[0]]['C1'] = 'Precision'
    workbook[workbook.sheetnames[0]]['D1'] = 'Recall'
    workbook[workbook.sheetnames[0]]['E1'] = 'Support'

    workbook[workbook.sheetnames[1]]['B1'] = 'Accuracy'
    workbook[workbook.sheetnames[1]]['C1'] = 'Precision'
    workbook[workbook.sheetnames[1]]['D1'] = 'Recall'
    workbook[workbook.sheetnames[1]]['E1'] = 'Support'
    # Save the workbook to a file
    workbook.save(filename)


def generate_summary(model_path):
    """
    generate model summary and save it an excel sheet

    Parameters:
        None

    Returns:
        None
    """
    models_list = []
    for model in glob.glob(model_path + '*.pt'):
        models_list.append(model)

    model = models.get_model(num_classes)
    device = torch.device('cuda:0')
    model = nn.DataParallel(model)
    model = model.cuda()
    model = model.to(device)

    prepare_excel(RESULTS_FILE)
    wb = openpyxl.load_workbook(RESULTS_FILE)
    i = 1
    shift = 1
    for name in models_list:
        if "Client" not in name:
            continue
        test_ds, val_ds = get_dataset(name)

        valid_loader = torch.utils.data.DataLoader(val_ds, batch_size=batchSize, shuffle='False', num_workers=0,
                                                   pin_memory=True)
        test_loader = torch.utils.data.DataLoader(test_ds, batch_size=batchSize, shuffle='False', num_workers=0,
                                                  pin_memory=True)

        ws = wb[wb.sheetnames[0]]

        model.load_state_dict(torch.load(name))
        # for i in range(5):
        model.eval()

        y, predictions = predict(model, valid_loader)
        Acc, wPr, wR, wF = calculate_scores(y, predictions)
        cl = 'A' + str(i + shift)
        ws[cl] = name
        cl = 'B' + str(i + shift)
        ws[cl] = Acc
        cl = 'C' + str(i + shift)
        ws[cl] = wF
        cl = 'D' + str(i + shift)
        ws[cl] = wPr
        cl = 'E' + str(i + shift)
        ws[cl] = wR
        cl = 'F' + str(i + shift)

        ws = wb[wb.sheetnames[1]]
        y, predictions = predict(model, test_loader)
        Acc, wPr, wR, wF = calculate_scores(y, predictions)

        cl = 'A' + str(i + shift)
        ws[cl] = name
        cl = 'B' + str(i + shift)
        ws[cl] = Acc
        cl = 'C' + str(i + shift)
        ws[cl] = wF
        cl = 'D' + str(i + shift)
        ws[cl] = wPr
        cl = 'E' + str(i + shift)
        ws[cl] = wR
        cl = 'F' + str(i + shift)
        i += 1

    wb.save(RESULTS_FILE)


def get_dataset(name):
    cid = -1
    if 'Client0' in name:
        cid = 0
    elif 'Client1' in name:
        cid = 1
    elif 'Client2' in name:
        cid = 2
    elif 'Client3' in name:
        cid = 3
    elif 'Client4' in name:
        cid = 4
    elif 'Client5' in name:
        cid = 5
    elif 'Client6' in name:
        cid = 6
    elif 'Client7' in name:
        cid = 7
    elif 'Client8' in name:
        cid = 8
    elif 'Client9' in name:
        cid = 9

    images, labels, test, validation = DATASET.get_global_test_data()

    img_test = images[test]
    lbl_test = labels[test]

    img_validation = images[validation]
    lbl_validation = labels[validation]

    test_ds = MyDataset(img_test, lbl_test, val_transform)
    val_ds = MyDataset(img_validation, lbl_validation, val_transform)

    return test_ds, val_ds


if __name__ == '__main__':
    generate_summary('../../models/')
